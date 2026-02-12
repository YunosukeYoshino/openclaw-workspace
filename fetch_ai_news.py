#!/usr/bin/env python3
"""
Yahoo! NewsからAI関連のニュースを取得してExcelにまとめる
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re

# AI関連のキーワード
AI_KEYWORDS = [
    'AI', '人工知能', 'ChatGPT', 'Claude', 'GPT', 'LLM', '機械学習',
    '深層学習', '生成AI', 'OpenAI', 'Anthropic', 'Google', 'Microsoft',
    'AI開発', 'AI技術', 'AIサービス', 'AIツール', 'AIモデル'
]

def fetch_yahoo_ai_news():
    """Yahoo! ニュースからAI関連記事を取得"""
    news_list = []

    try:
        # Yahoo! ニュースのトピックページ（AI・技術）
        urls = [
            'https://news.yahoo.co.jp/categories/it',
            'https://news.yahoo.co.jp/explore',
        ]

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

        for url in urls:
            try:
                response = requests.get(url, headers=headers, timeout=10)
                response.encoding = 'utf-8'

                soup = BeautifulSoup(response.text, 'html.parser')

                # ニュース記事のリンクを取得
                news_items = soup.find_all(['a'], class_=re.compile(r'newsFeed_item_link|sc-dzZBZM|sc-bfYlKJ'))

                for item in news_items:
                    title = item.get_text(strip=True)
                    link = item.get('href', '')

                    # AIキーワードが含まれる記事のみフィルタリング
                    if any(keyword.lower() in title.lower() for keyword in AI_KEYWORDS):
                        # 重複チェック
                        if not any(n['title'] == title for n in news_list):
                            news_list.append({
                                'title': title,
                                'link': link,
                                'source': 'Yahoo! News',
                                'date': datetime.now().strftime('%Y-%m-%d %H:%M')
                            })
            except Exception as e:
                print(f"Error fetching {url}: {e}")
                continue

        # 記事が少ない場合、トピック検索も試す
        if len(news_list) < 5:
            search_url = 'https://news.yahoo.co.jp/search?q=AI&ei=UTF-8'
            try:
                response = requests.get(search_url, headers=headers, timeout=10)
                response.encoding = 'utf-8'
                soup = BeautifulSoup(response.text, 'html.parser')

                # 検索結果の記事を取得
                search_items = soup.find_all(['a'], class_=re.compile(r'newsFeed_item'))

                for item in search_items:
                    title_elem = item.find('h1') or item.find('h2') or item.find('h3')
                    if title_elem:
                        title = title_elem.get_text(strip=True)
                        link_elem = item.find('a')
                        link = link_elem.get('href', '') if link_elem else ''

                        if title and not any(n['title'] == title for n in news_list):
                            news_list.append({
                                'title': title,
                                'link': link,
                                'source': 'Yahoo! News Search',
                                'date': datetime.now().strftime('%Y-%m-%d %H:%M')
                            })
            except Exception as e:
                print(f"Error searching: {e}")

    except Exception as e:
        print(f"Error in fetch_yahoo_ai_news: {e}")

    return news_list

def create_excel(news_list, filename='ai_news_report.xlsx'):
    """Excelファイルを作成"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AI ニュース'

    # ヘッダー
    headers = ['No.', 'タイトル', 'リンク', 'ソース', '取得日時']
    ws.append(headers)

    # ヘッダースタイル
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # データ行
    for idx, news in enumerate(news_list[:20], start=1):  # 最大20件
        ws.append([idx, news['title'], news['link'], news['source'], news['date']])

    # 列幅調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    # タイトル列の折り返し
    for row in ws.iter_rows(min_row=2, max_row=len(news_list)+1):
        row[1].alignment = Alignment(wrap_text=True, vertical='top')

    # タイトルシート（サマリー）
    if len(news_list) > 0:
        summary_ws = wb.create_sheet('サマリー', 0)

        summary_ws['A1'] = 'AI 最新ニュースレポート'
        summary_ws['A1'].font = Font(size=20, bold=True)
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws.merge_cells('A1:B1')

        summary_ws['A3'] = '作成日時:'
        summary_ws['B3'] = datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')

        summary_ws['A4'] = '記事数:'
        summary_ws['B4'] = len(news_list)

        summary_ws['A6'] = 'ソース:'
        summary_ws['B6'] = 'Yahoo! News'

        # 記事リスト
        summary_ws['A8'] = '記事一覧:'
        summary_ws['A8'].font = Font(bold=True)

        for idx, news in enumerate(news_list[:10], start=9):
            summary_ws[f'A{idx}'] = f'{idx-8}. {news["title"]}'

    wb.save(filename)
    print(f"Excelファイルを作成しました: {filename}")
    return filename

def main():
    print("Yahoo! ニュースからAI関連記事を取得中...")
    news_list = fetch_yahoo_ai_news()

    print(f"\n取得した記事数: {len(news_list)}件")

    if news_list:
        print("\n記事一覧:")
        for idx, news in enumerate(news_list, start=1):
            print(f"{idx}. {news['title'][:60]}...")

        filename = create_excel(news_list)
        return filename
    else:
        print("記事が取得できませんでした。")
        return None

if __name__ == '__main__':
    main()
