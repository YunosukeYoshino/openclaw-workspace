from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

news_data = [
    ["1", "AIが作業を代行してくれる「Claude Cowork」、Windows版が登場", "PC Watch", "2/12(木) 17:06", "https://news.yahoo.co.jp/articles/c9b4ef55dea56a5cb28751ec8077ef73c7b1a424"],
    ["2", "Coinbase、AIエージェント専用ウォレットインフラをリリース", "NADA NEWS", "2/12(木) 16:55", "https://news.yahoo.co.jp/articles/55f63bf61fa70ea18dfbfbc82ee1a1e77f323e21"],
    ["3", "「雲」から降りてきたAIは「パーソナル」な存在になれるのか――開催から1カ月経過した「CES 2026」を振り返る", "ITmedia PC USER", "2/12(木) 17:05", "https://news.yahoo.co.jp/articles/5ad407ec4ccebd50cf11e1a32a94f1ef7bc85034"],
    ["4", "AIで週1～7時間の業務を削減--AIが生成した成果物の修正や検証に週1～4時間", "ZDNET Japan", "2/12(木) 16:30", "https://news.yahoo.co.jp/articles/6d3b90872233cba6a04b265ad522d2426e67fd41"],
    ["5", "セーフィー、映像データとAIを組み合わせた「Safie AI Studio」を提供", "ZDNET Japan", "2/12(木) 15:51", "https://news.yahoo.co.jp/articles/1f3c617361e824e9da927d9e4d0ede7ee7e3f3c0"],
    ["6", "「高市首相vs.ウルトラマン」「悟空vs.ドラえもん」も……中国発の新動画AI「Seedance2.0」物議", "ITmedia NEWS", "2/12(木) 12:16", "https://news.yahoo.co.jp/articles/0165c53be5f2d19804bc8a157c369b4027dda786"],
    ["7", "AI要約でサイト「素通り」傾向は", "Yahoo! News", "-", "https://news.yahoo.co.jp/pickup/6569394"],
]

wb = Workbook()
ws = wb.active
ws.title = "AI 最新ニュース"

headers = ["No.", "タイトル", "ソース", "日時", "URL"]
ws.append(headers)

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=12, name='Arial')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

title_font = Font(size=11, name='Arial')
title_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
data_font = Font(size=10, name='Arial')
data_alignment = Alignment(horizontal='left', vertical='center')

for row_data in news_data:
    ws.append(row_data)

for row in ws.iter_rows(min_row=2, max_row=len(news_data)+1):
    row[0].font = data_font
    row[0].alignment = data_alignment
    row[0].border = border
    
    row[1].font = title_font
    row[1].alignment = title_alignment
    row[1].border = border
    
    row[2].font = data_font
    row[2].alignment = data_alignment
    row[2].border = border
    
    row[3].font = data_font
    row[3].alignment = Alignment(horizontal='center')
    row[3].border = border
    
    row[4].font = Font(size=9, name='Arial', color='0070C0', underline='single')
    row[4].alignment = data_alignment
    row[4].border = border

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 50

summary_ws = wb.create_sheet('サマリー', 0)

summary_ws['A1'] = 'AI 最新ニュースレポート'
summary_ws['A1'].font = Font(size=18, bold=True, name='Arial', color='4472C4')
summary_ws['A1'].alignment = Alignment(horizontal='center')
summary_ws.merge_cells('A1:B1')

summary_ws['A3'] = '作成日時:'
summary_ws['A3'].font = Font(bold=True, size=11, name='Arial')
summary_ws['B3'] = datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')
summary_ws['B3'].font = Font(size=11, name='Arial')

summary_ws['A4'] = 'ニュース数:'
summary_ws['A4'].font = Font(bold=True, size=11, name='Arial')
summary_ws['B4'] = len(news_data)
summary_ws['B4'].font = Font(bold=True, size=11, name='Arial')

summary_ws['A6'] = 'ソース:'
summary_ws['A6'].font = Font(bold=True, size=11, name='Arial')
summary_ws['B6'] = 'Yahoo! News (ITカテゴリ)'

summary_ws['A8'] = 'カテゴリ:'
summary_ws['A8'].font = Font(bold=True, size=11, name='Arial')
summary_ws['B8'] = 'AI / 人工知能 / 生成AI'

sources = set(n[2] for n in news_data)
summary_ws['A10'] = '掲載媒体一覧:'
summary_ws['A10'].font = Font(bold=True, size=11, name='Arial')

for idx, source in enumerate(sources, start=11):
    summary_ws[f'A{idx}'] = f'  • {source}'
    summary_ws[f'A{idx}'].font = Font(size=10, name='Arial')

wb.save('/workspace/ai_news_report.xlsx')
print('Excelファイルを作成しました: ai_news_report.xlsx')
